from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd
import os

from Journal import all_journals
from Ledger import get_ledger_format
from Trial_Balance import get_trial_balance
####################################################################################################################################################################################

def get_export_path(filename, custom_path=None):
    downloads_folder = os.path.join(os.path.expanduser("~"), "Documents")
    folder = custom_path if custom_path else downloads_folder
    return os.path.join(folder, filename)


def export_journal_xlsx():
    journals = all_journals()
    
    if not journals:
        return False, "No journal entries to export."

    # Create a DataFrame
    columns = ["Index", "Date & Time", "Debit Account", "Credit Account", "Amount", "Narration"]
    df = pd.DataFrame(journals, columns=columns)

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Entries"

    # Write the DataFrame to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Style the header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

    # Style the data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    save_path = get_export_path("Journal.xlsx")

    try:
        wb.save(save_path)
        return True, f"Journal entries exported successfully to {save_path}", save_path
    except PermissionError:
        return False, f"Permission denied: Unable to save file to {save_path}", None    
    except OSError as e:
        return False, f"Error saving file: {e}", None

def export_ledger_xlsx():
    ledger_data = get_ledger_format()
    
    if not ledger_data:
        return False, "No ledger data to export.", None

    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    for account in ledger_data.keys():
        # Create a new sheet for each account
        ws = wb.create_sheet(title=account)
        for r in dataframe_to_rows(ledger_data[account], index=False, header=True):
            ws.append(r)

        # Style the header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

        # Style the data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    save_path = get_export_path("Ledger.xlsx")

    try:
        wb.save(save_path)
        return True, f"Ledger exported successfully to {save_path}", save_path
    except PermissionError:
        return False, f"Permission denied: Unable to save file to {save_path}.", None
    except OSError as e:
        return False, f"Error saving file: {e}", None

def export_tb_xlsx():
    trial_balance_data = get_trial_balance()
    if not trial_balance_data:
        return False, "No trial balance data to export.", None

    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    headers = ['Particulars', 'Debit', 'Credit']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row = 1, column = col, value = header)
        ws.column_dimensions[get_column_letter(col)].width = 15
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.column_dimensions['A'].width = 35
    
    row = 2
    debit_total = 0
    credit_total = 0
    
    for account, debit, credit in trial_balance_data:
        ws.cell(row=row, column=1, value=account)
        ws.cell(row=row, column=2, value=debit)        
        ws.cell(row=row, column=3, value=credit)
        debit_total += debit
        credit_total += credit
        row += 1
    
    row += 1
    ws.cell(row=row, column=1, value="Total")
    ws.cell(row=row, column=2, value=debit_total)
    ws.cell(row=row, column=3, value=credit_total)

    # Style the header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

    # Style the data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    # Style the total row
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    save_path = get_export_path("Trial Balance.xlsx")

    try:
        wb.save(save_path)
        return True, f"Trial Balance exported successfully to {save_path}", save_path
    except PermissionError:
        return False, f"Permission denied: Unable to save file to {save_path}.", None
    except OSError as e:
        return False, f"Error saving file: {e}", None

def export_all_xlsx():
    journal_success, journal_message = export_journal_xlsx()
    ledger_success, ledger_message = export_ledger_xlsx()
    tb_success, tb_message = export_tb_xlsx()
    
    if journal_success and ledger_success and tb_success:
        return True, "All data exported successfully."
    else:
        failed_exports = []
        if not journal_success:
            failed_exports.append("Journal")
        if not ledger_success:
            failed_exports.append("Ledger")
        if not tb_success:
            failed_exports.append("Trial Balance")
        return False, f"Failed to export: {', '.join(failed_exports)}"
